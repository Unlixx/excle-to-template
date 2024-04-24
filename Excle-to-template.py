# python3
# Excle-to-template.py
# 指定文件转换到模板

import tkinter as tk
from tkinter import filedialog, Text, Scrollbar
import sys
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import re
from datetime import datetime
import shutil


class App:

    def __init__(self, master):
        self.master = master
        master.title("模板转换")

        # 创建用于显示print内容的文本框和滚动条
        self.output_text = Text(master, height=10, width=50)
        self.output_text.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

        scrollbar = Scrollbar(command=self.output_text.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.output_text.config(yscrollcommand=scrollbar.set)

        # 重定向标准输出到Text组件
        sys.stdout = PrintToText(self.output_text)

        # 在这里添加默认的文字信息
        print("""欢迎使用模板转换程序！（内部使用）
        注意事项：
        1.点运行前请暂时关闭360等杀毒软件。"
        2.原始文件请转换为xlsx后缀。""")

        # 创建源文件选择的标签和按钮
        self.source_label = tk.Label(master, text="原始文件（请再次确认原始文件的后缀为“xlsx”）")
        self.source_label.pack()

        self.source_button = tk.Button(master, text="选择文件", command=self.select_source_file)
        self.source_button.pack()

        # 创建模板文件选择的标签和按钮
        self.template_label = tk.Label(master, text="模板文件")
        self.template_label.pack()

        self.template_button = tk.Button(master, text="选择文件", command=self.select_template_file)
        self.template_button.pack()

        # 创建继续执行的按钮
        self.continue_button = tk.Button(master, text="运行", command=self.continue_execution)
        self.continue_button.pack()

    def select_source_file(self):
        file_path = filedialog.askopenfilename(title="Select Source File")
        if file_path:
            self.source_label.config(text="Source File: " + file_path)

    def select_template_file(self):
        template_path = filedialog.askopenfilename(title="Select Template File")
        if template_path:
            self.template_label.config(text="Template File: " + template_path)

    def continue_execution(self):
        # 保存原始工作目录
        original_work_dir = os.getcwd()
        # 在这里添加剩余代码逻辑
        print("执行转换中...")

        # 获取源文件和模板文件的路径
        source_file_path = self.source_label["text"].split(": ")[1].strip()
        template_file_path = self.template_label["text"].split(": ")[1].strip()

        # 确保路径不为空
        if not source_file_path or not template_file_path:
            print("请先选择源文件和模板文件。")
            return
        # 设置新文件夹的名称
        new_folder = "拆分文件"

        # 确保新文件夹存在
        if not os.path.exists(new_folder):
            os.makedirs(new_folder)

        # 载入原始的Excel文件
        file_path = source_file_path
        wb = load_workbook(file_path)
        ws = wb.active

        # 初始化起始行和拆分行数模式
        start = 1
        mode = True  # 使用布尔值来切换拆分行数模式

        # 循环直到下一个循环内的所有行没有内容
        while True:
            if mode:
                end = start + 14 - 1  # 计算结束行
            else:
                end = start + 13 - 1  # 计算结束行

            # 检查结束行是否超出了Excel文件的最大行数
            if end > ws.max_row:
                break

            # 创建新的工作簿并复制指定范围内的行
            wb_new = Workbook()  # 创建一个新的空工作簿
            ws_new = wb_new.active  # 获取活动的工作表

            for row in ws.iter_rows(min_row=start, max_row=end, values_only=True):
                ws_new.append(row)

            # 设置新的文件名和路径
            output_file = os.path.join(new_folder, f'{start}-{end}.xlsx')

            # 保存新的Excel文件到新文件夹
            wb_new.save(output_file)
            print(f'已保存文件：{output_file}')

            # 更新起始行为下一个循环的起始行
            start = end + 1

            # 切换拆分行数模式
            mode = not mode

        print('拆分完成！')

        # 设置文件夹路径，使用原始字符串
        folder_path = r'拆分文件'

        # 获取指定文件夹下所有Excel文件的路径
        excel_files = [os.path.join(folder_path, filename) for filename in os.listdir(folder_path)
                       if filename.endswith('.xlsx')]

        # 定义一个函数来检查并设置默认值
        def set_default_values(df, col1, col2):
            # 确保列名在DataFrame中存在
            if col1 in df.columns and col2 in df.columns:
                # 如果col1列有值且为数字，col2列没有值，则将col2列的默认值设置为0
                df.loc[(df[col1].notnull() & pd.to_numeric(df[col1], errors='coerce').notnull()) &
                       (df[col2].isnull()), col2] = 0
                # 如果col2列有值且为数字，col1列没有值，则将col1列的默认值设置为0
                df.loc[(df[col2].notnull() & pd.to_numeric(df[col2], errors='coerce').notnull()) &
                       (df[col1].isnull()), col1] = 0
            return df

        # 遍历指定文件夹下所有Excel文件
        for excel_file in excel_files:
            # 读取当前Excel文件
            df = pd.read_excel(excel_file)

            # 检查列名是否存在
            if 'Unnamed: 8' in df.columns and 'Unnamed: 10' in df.columns:
                # 调用函数设置默认值
                df = set_default_values(df, 'Unnamed: 8', 'Unnamed: 10')

                # 保存修改后的DataFrame回原文件
                df.to_excel(excel_file, index=False)
                print(f"在{excel_file}中需要添加0的列添加0成功。")
            else:
                print(f"在{excel_file}中，需要添加0的列名没找到。")

        source_folder_path = r'拆分文件'  # 源文件夹路径
        destination_folder_path = r'复制到模板'  # 目标文件夹路径

        # 如果目标文件夹不存在，则创建它
        if not os.path.exists(destination_folder_path):
            os.makedirs(destination_folder_path)

        # 获取源文件夹下所有的Excel文件
        excel_files = [f for f in os.listdir(source_folder_path) if f.endswith('.xlsx')]

        # 遍历所有的Excel文件
        for source_file_name in excel_files:
            # 构造源文件的完整路径
            source_file_path = os.path.join(source_folder_path, source_file_name)

            # 构造目标文件名，并设置目标文件的完整路径
            new_destination_file_name = f'new_{source_file_name}.xlsx'
            destination_file_path = os.path.join(destination_folder_path, new_destination_file_name)

            # 加载源Excel文件
            source_workbook = load_workbook(source_file_path)
            source_sheet = source_workbook.active  # 假设数据在第一个工作表

            # 从源Excel文件的A7至A11单元格读取内容
            data_to_copy_a = [cell.value for row in source_sheet['A7:A11'] for cell in row if cell.value is not None]

            # 从源Excel文件的D7至C11单元格读取内容，并只提取数字
            data_to_copy_c = [(re.sub(r'[^\d]', '', str(cell.value)))
                              if cell.value is not None else None for row in source_sheet['D7:D11'] for cell in row]

            # 从源Excel文件的I7至D11单元格读取内容
            data_to_copy_d = [cell.value for row in source_sheet['I7:I11'] for cell in row if cell.value is not None]

            # 从源Excel文件的K7至K11单元格读取内容
            data_to_copy_e = [cell.value for row in source_sheet['K7:K11'] for cell in row if cell.value is not None]

            # 从源Excel文件的E3单元格读取内容
            if source_sheet['E3'].value is not None:
                data_to_copy_time = source_sheet['E3'].value
                data_to_copy_time = datetime.strptime(data_to_copy_time, "%Y年%m月%d日")
                data_to_copy_time = data_to_copy_time.strftime("%Y-%m-%d")

            else:
                data_to_copy_time = None  # 或者你可以设置一个默认值

            # 从源Excel文件的I5单元格读取内容
            if source_sheet['I5'].value is not None:
                # 读取I5单元格的内容
                d4_value = source_sheet['I5'].value

                # 使用正则表达式提取数字，这里假设value是一个字符串
                # 如果value不是字符串，需要先转换为字符串，如：str(value)
                data_to_copy_d4 = re.sub(r'[^\d]', '', str(d4_value))

                # 现在 data_to_copy_time 只包含原字符串中的数字

            # 从源Excel文件的J4单元格读取内容
            if source_sheet['J4'].value is not None:
                # 读取D4单元格的内容
                j4_value = source_sheet['J4'].value

            # 加载目标Excel文件，如果是新文件则创建一个
            if os.path.exists(destination_file_path):
                destination_workbook = load_workbook(destination_file_path)
            else:
                destination_workbook = load_workbook(template_file_path)
                destination_sheet = destination_workbook.active  # 假设我们要写入第一个工作表

            # 将源Excel文件的A7至A11单元格内容复制到目标Excel文件的B2至B6单元格
            for index, value in enumerate(data_to_copy_a, start=1):
                destination_sheet.cell(row=index + 1, column=2).value = value

            # 将源Excel文件的D7至C11单元格内容复制到目标Excel文件的C2至C6单元格
            for index, value in enumerate(data_to_copy_c, start=1):
                if value is not None:
                    destination_sheet.cell(row=index + 1, column=3).value = value

            # 将源Excel文件的I7至D11单元格内容复制到目标Excel文件的D2至D6单元格
            for index, value in enumerate(data_to_copy_d, start=1):
                destination_sheet.cell(row=index + 1, column=4).value = value

            # 将源Excel文件的K7至K11单元格内容复制到目标Excel文件的E2至E6单元格
            for index, value in enumerate(data_to_copy_e, start=1):
                destination_sheet.cell(row=index + 1, column=5).value = value

            # 将源Excel文件的E3单元格内容复制到目标Excel文件的G2-G6单元格
            if data_to_copy_time is not None:
                if destination_sheet['B2'].value is not None:
                    destination_sheet['G2'] = data_to_copy_time
                if destination_sheet['B3'].value is not None:
                    destination_sheet['G3'] = data_to_copy_time
                if destination_sheet['B4'].value is not None:
                    destination_sheet['G4'] = data_to_copy_time
                if destination_sheet['B5'].value is not None:
                    destination_sheet['G5'] = data_to_copy_time
                if destination_sheet['B6'].value is not None:
                    destination_sheet['G6'] = data_to_copy_time

            # 将源Excel文件的I5单元格内容复制到目标Excel文件的H2-H6单元格
            if data_to_copy_d4 is not None:
                if destination_sheet['B2'].value is not None:
                    destination_sheet['H2'] = data_to_copy_d4
                if destination_sheet['B3'].value is not None:
                    destination_sheet['H3'] = data_to_copy_d4
                if destination_sheet['B4'].value is not None:
                    destination_sheet['H4'] = data_to_copy_d4
                if destination_sheet['B5'].value is not None:
                    destination_sheet['H5'] = data_to_copy_d4
                if destination_sheet['B6'].value is not None:
                    destination_sheet['H6'] = data_to_copy_d4

                # 将源Excel文件的J4单元格内容复制到目标Excel文件的H2-H6单元格
                if j4_value is not None:
                    if destination_sheet['B2'].value is not None:
                        destination_sheet['A2'] = j4_value
                    if destination_sheet['B3'].value is not None:
                        destination_sheet['A3'] = j4_value
                    if destination_sheet['B4'].value is not None:
                        destination_sheet['A4'] = j4_value
                    if destination_sheet['B5'].value is not None:
                        destination_sheet['A5'] = j4_value
                    if destination_sheet['B6'].value is not None:
                        destination_sheet['A6'] = j4_value

                # 将修改后的目标Excel文件另存为新文件到目标文件夹
                destination_workbook.save(destination_file_path)

                # 关闭源Excel文件
                source_workbook.close()
                # 关闭目标Excel文件
                destination_workbook.close()

            print(f"正在将原始表格数据都复制到模板中。")

        # 设置工作目录到指定的文件夹
        os.chdir(r'复制到模板')

        # 获取当前目录下所有的Excel文件，并对它们进行排序以确保顺序
        files = sorted([f for f in os.listdir() if f.endswith('.xlsx')],
                       key=lambda x: int(re.search(r'\d+', os.path.splitext(x)[0]).group()))

        # 初始化一个空的DataFrame用于存放合并后的数据
        all_data_h = pd.DataFrame()

        # 初始化一个列表，用于存放每个文件的数据
        data_frames = []

        # 遍历文件列表，逐个读取文件内容
        for file in files:
            current_data = pd.read_excel(file)
            # 将每个文件的数据存储在列表中，而不是直接合并
            data_frames.append(current_data)

        # 使用pd.concat一次性合并列表中的所有DataFrame
        all_data_h = pd.concat(data_frames, ignore_index=True)

        # 获取当前工作目录
        current_work_dir = os.getcwd()

        # 获取上一级目录的路径
        parent_dir = os.path.dirname(current_work_dir)

        # 将合并后的数据保存到上一级目录中的新Excel文件中
        output_filename = '合并后的文件.xlsx'
        output_path = os.path.join(parent_dir, output_filename)

        # 将合并后的数据保存到名为‘合并后的文件.xlsx’的文件中，不包含索引
        all_data_h.to_excel(output_path, index=False)

        # 将工作目录改回原始工作目录
        os.chdir(original_work_dir)

        # 获取程序运行的目录（当前工作目录）
        current_dir = os.path.dirname(sys.argv[0])

        # 定义要删除的文件夹名称列表
        folders_to_delete = ["拆分文件", "复制到模板"]

        # 循环删除每个文件夹
        for folder_name in folders_to_delete:
            # 构建目标文件夹的完整路径
            target_folder_path = os.path.join(current_dir, folder_name)

            # 检查文件夹是否存在
            if os.path.exists(target_folder_path):
                try:
                    # 尝试删除文件夹
                    shutil.rmtree(target_folder_path)
                    print(f"已删除文件夹：{target_folder_path}")
                except Exception as e:
                    # 如果发生错误，打印出错误信息
                    print(f"删除文件夹时出错：{target_folder_path}, 错误信息：{e}")
            else:
                print(f"文件夹不存在，无需删除：{target_folder_path}")

        print("完成！完成的文件名为‘合并后的文件.xlsx’"
              "点击右上角的×退出程序")


class PrintToText:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, s):
        self.text_widget.configure(state='normal')
        self.text_widget.insert('end', s + '\n')
        self.text_widget.see('end')
        self.text_widget.update()
        self.text_widget.update_idletasks()  # 添加这一行
        self.text_widget.configure(state='disabled')

    def flush(self):  # 添加flush方法
        pass


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
